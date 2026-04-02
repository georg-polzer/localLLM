import asyncio
import contextlib
import json
import os
import re
import urllib.error
import urllib.request
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from uuid import uuid4

from fastapi import FastAPI, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from starlette.requests import Request

try:
    import mcp.types as mcp_types
    from mcp.client.session import ClientSession
    from mcp.client.streamable_http import streamable_http_client
except ImportError:  # pragma: no cover
    mcp_types = None
    ClientSession = None
    streamable_http_client = None


BASE_DIR = Path(__file__).parent
app = FastAPI(title="Local LLM Chat")
app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))
OLLAMA_BASE_URL = os.environ.get("OLLAMA_BASE_URL", "http://127.0.0.1:11434").rstrip("/")
OLLAMA_MODEL = os.environ.get("OLLAMA_MODEL", "gemma4:e4b")
DUODATA_ENV_PATH = Path("/Users/georgpolzer/Documents/DuoData/DuoData MCP/.env")
DEMO_WORKBOOKS = {
    "pharma_demo_metrics.xlsx": Path(
        "/Users/georgpolzer/Documents/DuoData/DuoData MCP/demo/workbooks/pharma_demo_metrics.xlsx"
    ),
    "public_transport_demo_metrics.xlsx": Path(
        "/Users/georgpolzer/Documents/DuoData/DuoData MCP/demo/workbooks/public_transport_demo_metrics.xlsx"
    ),
}

AGENT_MAX_STEPS = 8
AGENT_DUODATA_TOOL_NAMES = {"list_metrics", "get_metric", "resolve_metric", "plan_metric_execution"}
DUODATA_API_BASE_URL = os.environ.get("DUODATA_API_BASE_URL", "http://127.0.0.1:8002").rstrip("/")


def normalize_endpoint(endpoint: str) -> str:
    return endpoint.strip().rstrip("/")


def make_config_signature(system_prompt: str | None, mcp_endpoints: list[str]) -> str:
    normalized_endpoints = [normalize_endpoint(endpoint) for endpoint in mcp_endpoints]
    return json.dumps(
        {
            "system_prompt": system_prompt or "",
            "mcp_endpoints": normalized_endpoints,
        },
        sort_keys=True,
    )


def compact_json(data: Any) -> str:
    return json.dumps(data, ensure_ascii=True, separators=(",", ":"), sort_keys=True)


def normalize_metric_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def schema_to_prompt(schema: dict[str, Any] | None) -> str:
    if not schema:
        return "Provide an empty JSON object: {}"

    payload = compact_json(schema)
    if len(payload) > 1400:
        payload = payload[:1400] + "...(truncated)"
    return f"Provide a JSON object that matches this schema: {payload}"


def parse_json_object(raw: str) -> dict[str, Any]:
    candidate = raw.strip()
    if candidate.startswith("```"):
        candidate = re.sub(r"^```(?:json)?\s*", "", candidate)
        candidate = re.sub(r"\s*```$", "", candidate)

    if not candidate:
        return {}

    parsed = json.loads(candidate)
    if not isinstance(parsed, dict):
        raise ValueError("Tool arguments must decode to a JSON object.")
    return parsed


def render_tool_content(block: Any) -> str:
    if isinstance(block, getattr(mcp_types, "TextContent", ())):
        return block.text

    if isinstance(block, getattr(mcp_types, "ImageContent", ())):
        return f"[Image content omitted: {block.mimeType}]"

    if isinstance(block, getattr(mcp_types, "AudioContent", ())):
        return f"[Audio content omitted: {block.mimeType}]"

    if isinstance(block, getattr(mcp_types, "EmbeddedResource", ())):
        resource = block.resource
        text = getattr(resource, "text", None)
        if text:
            return f"[Embedded resource]\n{text}"
        blob = getattr(resource, "blob", None)
        mime_type = getattr(resource, "mimeType", "application/octet-stream")
        return f"[Embedded resource omitted: {mime_type}, {len(blob or '')} bytes]"

    if isinstance(block, getattr(mcp_types, "ResourceLink", ())):
        label = block.title or block.name
        return f"[Resource link] {label}: {block.uri}"

    return compact_json(block.model_dump(mode="json"))


async def list_all_tools(client: "ClientSession") -> list["mcp_types.Tool"]:
    tools: list[mcp_types.Tool] = []
    cursor: str | None = None

    while True:
        response = await client.list_tools(cursor=cursor) if cursor else await client.list_tools()
        tools.extend(response.tools)
        if not response.nextCursor:
            break
        cursor = response.nextCursor

    return tools


class ChatRequest(BaseModel):
    message: str = Field(..., min_length=1)
    conversation_id: str | None = None
    system_prompt: str | None = None
    mcp_endpoints: list[str] = Field(default_factory=list)


class ResetRequest(BaseModel):
    conversation_id: str


class MCPInspectRequest(BaseModel):
    endpoints: list[str] = Field(default_factory=list)


@dataclass
class EndpointConnection:
    endpoint: str
    client: "ClientSession"
    tools: list["mcp_types.Tool"]
    lock: asyncio.Lock


@dataclass
class ManagedSession:
    signature: str
    history: list[dict[str, str]]
    mcp_bridge: "MCPBridge | None" = None

    async def close(self) -> None:
        if self.mcp_bridge is not None:
            await self.mcp_bridge.close()


class MCPBridge:
    def __init__(self, stack: contextlib.AsyncExitStack, endpoints: list[EndpointConnection]) -> None:
        self._stack = stack
        self._endpoints = {connection.endpoint: connection for connection in endpoints}

    @classmethod
    async def create(cls, endpoints: list[str]) -> "MCPBridge":
        if streamable_http_client is None or ClientSession is None or mcp_types is None:
            raise HTTPException(
                status_code=500,
                detail="MCP support is unavailable. Install the 'mcp' Python package first.",
            )

        normalized = [normalize_endpoint(endpoint) for endpoint in endpoints if normalize_endpoint(endpoint)]
        if not normalized:
            raise HTTPException(status_code=400, detail="At least one MCP endpoint is required.")

        stack = contextlib.AsyncExitStack()
        connections: list[EndpointConnection] = []

        try:
            for endpoint in normalized:
                read_stream, write_stream, _ = await stack.enter_async_context(
                    streamable_http_client(endpoint)
                )
                client = ClientSession(read_stream, write_stream)
                await stack.enter_async_context(client)
                await client.initialize()
                tools = await list_all_tools(client)
                connections.append(
                    EndpointConnection(
                        endpoint=endpoint,
                        client=client,
                        tools=tools,
                        lock=asyncio.Lock(),
                    )
                )
            return cls(stack, connections)
        except Exception:
            await stack.aclose()
            raise

    def summaries(self) -> list[dict[str, Any]]:
        summaries = []
        for connection in self._endpoints.values():
            summaries.append(
                {
                    "endpoint": connection.endpoint,
                    "tool_count": len(connection.tools),
                    "tools": [
                        {
                            "name": tool.name,
                            "title": tool.title,
                            "description": tool.description,
                        }
                        for tool in connection.tools
                    ],
                }
            )
        return summaries

    def tool_catalog(self) -> list[dict[str, Any]]:
        catalog: list[dict[str, Any]] = []
        for connection in self._endpoints.values():
            for tool in connection.tools:
                catalog.append(
                    {
                        "endpoint": connection.endpoint,
                        "name": tool.name,
                        "title": tool.title,
                        "description": tool.description,
                        "inputSchema": tool.inputSchema,
                    }
                )
        return catalog

    def endpoint_for_tool(self, tool_name: str) -> str | None:
        for connection in self._endpoints.values():
            if any(tool.name == tool_name for tool in connection.tools):
                return connection.endpoint
        return None

    async def call_tool(self, endpoint: str, name: str, arguments: dict[str, Any]) -> str:
        connection = self._endpoints[endpoint]

        async with connection.lock:
            result = await connection.client.call_tool(name=name, arguments=arguments)

        parts: list[str] = []
        if result.structuredContent is not None:
            parts.append("Structured result:\n" + json.dumps(result.structuredContent, indent=2, ensure_ascii=True))

        if result.content:
            rendered = [render_tool_content(block) for block in result.content]
            parts.append("\n\n".join(part for part in rendered if part))

        if not parts:
            parts.append("Tool completed without returning content.")

        final = "\n\n".join(parts)
        if result.isError:
            return f"Tool error from {name}:\n{final}"
        return final

    async def call_tool_structured(
        self, endpoint: str, name: str, arguments: dict[str, Any]
    ) -> tuple[dict[str, Any] | None, str]:
        connection = self._endpoints[endpoint]

        async with connection.lock:
            result = await connection.client.call_tool(name=name, arguments=arguments)

        text_parts = [render_tool_content(block) for block in result.content] if result.content else []
        text = "\n\n".join(part for part in text_parts if part)
        structured = result.structuredContent if isinstance(result.structuredContent, dict) else None

        if result.isError:
            detail = text or compact_json(structured or {})
            raise RuntimeError(f"Tool error from {name}: {detail}")

        return structured, text

    async def build_grounding_context(self, user_message: str) -> str | None:
        message_lower = user_message.lower()

        for connection in self._endpoints.values():
            available_tools = {tool.name for tool in connection.tools}
            if not {"list_metrics", "resolve_metric"}.issubset(available_tools):
                continue

            metrics_payload, metrics_text = await self.call_tool_structured(
                connection.endpoint, "list_metrics", {}
            )
            metrics = metrics_payload.get("metrics", []) if metrics_payload else []
            matched_metrics = [
                metric for metric in metrics if metric.get("name", "").lower() in message_lower
            ]

            context_blocks: list[str] = []

            if (
                "metric" in message_lower
                or "kpi" in message_lower
                or "source" in message_lower
                or "available" in message_lower
            ):
                context_blocks.append("DuoData metric catalog:")
                preview = metrics[:8]
                for metric in preview:
                    context_blocks.append(
                        f"- {metric.get('name')}: source={metric.get('source') or 'Derived / multi-source'}, period={metric.get('period')}"
                    )

            for metric in matched_metrics[:3]:
                metric_name = metric.get("name")
                if not metric_name:
                    continue

                resolved_payload, resolved_text = await self.call_tool_structured(
                    connection.endpoint,
                    "resolve_metric",
                    {"metric_name": metric_name},
                )
                context_blocks.append(f"DuoData resolved metric: {metric_name}")
                context_blocks.append(
                    compact_json(resolved_payload) if resolved_payload else resolved_text
                )

                if "plan_metric_execution" in available_tools:
                    planned_payload, planned_text = await self.call_tool_structured(
                        connection.endpoint,
                        "plan_metric_execution",
                        {"metric_name": metric_name, "prefer_backend": "snowflake"},
                    )
                    context_blocks.append(
                        compact_json(planned_payload) if planned_payload else planned_text
                    )

            if context_blocks:
                return "\n".join(context_blocks)

            if metrics_text:
                return metrics_text

        return None

    async def close(self) -> None:
        await self._stack.aclose()


class SessionStore:
    def __init__(self) -> None:
        self._lock = asyncio.Lock()
        self._sessions: dict[str, ManagedSession] = {}

    async def get_or_create(
        self,
        conversation_id: str,
        system_prompt: str | None = None,
        mcp_endpoints: list[str] | None = None,
    ) -> ManagedSession:
        normalized_endpoints = [normalize_endpoint(endpoint) for endpoint in (mcp_endpoints or []) if normalize_endpoint(endpoint)]
        signature = make_config_signature(system_prompt, normalized_endpoints)

        async with self._lock:
            existing = self._sessions.get(conversation_id)
            if existing and existing.signature == signature:
                return existing

            mcp_bridge = None
            if normalized_endpoints:
                mcp_bridge = await MCPBridge.create(normalized_endpoints)
            managed = ManagedSession(signature=signature, history=[], mcp_bridge=mcp_bridge)

            if existing is not None:
                await existing.close()

            self._sessions[conversation_id] = managed
            return managed

    async def reset(self, conversation_id: str) -> bool:
        async with self._lock:
            managed = self._sessions.pop(conversation_id, None)

        if managed is None:
            return False

        await managed.close()
        return True

    async def close_all(self) -> None:
        async with self._lock:
            sessions = list(self._sessions.values())
            self._sessions.clear()

        for managed in sessions:
            await managed.close()


store = SessionStore()


@app.on_event("shutdown")
async def shutdown_event() -> None:
    await store.close_all()


def _ollama_request(path: str, payload: dict[str, Any]) -> urllib.request.Request:
    return urllib.request.Request(
        f"{OLLAMA_BASE_URL}{path}",
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )


def _ollama_get(path: str) -> urllib.request.Request:
    return urllib.request.Request(f"{OLLAMA_BASE_URL}{path}", method="GET")


async def ollama_status() -> tuple[bool, str]:
    def _fetch() -> tuple[bool, str]:
        try:
            with urllib.request.urlopen(_ollama_get("/api/tags"), timeout=10) as response:
                payload = json.loads(response.read().decode("utf-8"))
        except urllib.error.URLError as exc:
            return False, f"Ollama unavailable: {exc.reason}"
        except Exception as exc:  # pragma: no cover
            return False, f"Ollama unavailable: {exc}"

        names = {model.get("name") for model in payload.get("models", [])}
        if OLLAMA_MODEL in names:
            return True, f"Ollama ready ({OLLAMA_MODEL})"
        return False, f"Ollama is running but model '{OLLAMA_MODEL}' is not installed"

    return await asyncio.to_thread(_fetch)


async def ollama_stream_chat(messages: list[dict[str, str]]):
    def _iter_lines():
        request = _ollama_request(
            "/api/chat",
            {"model": OLLAMA_MODEL, "messages": messages, "stream": True},
        )
        with urllib.request.urlopen(request, timeout=120) as response:
            for raw_line in response:
                line = raw_line.decode("utf-8").strip()
                if line:
                    yield line

    iterator = await asyncio.to_thread(lambda: iter(_iter_lines()))
    sentinel = object()

    def _next_or_sentinel():
        return next(iterator, sentinel)

    while True:
        line = await asyncio.to_thread(_next_or_sentinel)
        if line is sentinel:
            break
        yield json.loads(line)


async def ollama_complete(messages: list[dict[str, str]], *, json_mode: bool = False) -> str:
    def _fetch() -> str:
        payload: dict[str, Any] = {
            "model": OLLAMA_MODEL,
            "messages": messages,
            "stream": False,
            "options": {"temperature": 0},
        }
        if json_mode:
            payload["format"] = "json"
        request = _ollama_request("/api/chat", payload)
        with urllib.request.urlopen(request, timeout=120) as response:
            body = json.loads(response.read().decode("utf-8"))
        return body.get("message", {}).get("content", "")

    return await asyncio.to_thread(_fetch)


def load_duodata_env() -> dict[str, str]:
    values: dict[str, str] = {}
    for line in DUODATA_ENV_PATH.read_text().splitlines():
        if line and not line.startswith("#") and "=" in line:
            key, value = line.split("=", 1)
            values[key] = value
    return values


def snowflake_date_from_serial(raw: str) -> date:
    return date(1970, 1, 1) + timedelta(days=int(raw))


def execute_snowflake_sql(statement: str, *, database: str | None = None, schema: str | None = None) -> dict[str, Any]:
    normalized = statement.lstrip().lower()
    if not normalized.startswith(("select", "with")):
        raise ValueError("Only read-only SELECT or WITH statements are allowed.")
    env = load_duodata_env()
    payload: dict[str, Any] = {
        "statement": statement,
        "timeout": 60,
        "warehouse": env.get("DUODATA_SNOWFLAKE_SQL_API_WAREHOUSE"),
        "role": env.get("DUODATA_SNOWFLAKE_SQL_API_ROLE"),
    }
    if database:
        payload["database"] = database
    if schema:
        payload["schema"] = schema

    headers = {
        "Authorization": f"Bearer {env['DUODATA_SNOWFLAKE_BEARER_TOKEN']}",
        "Content-Type": "application/json",
        "Accept": "application/json",
        "User-Agent": "local-llm-chat/0.1.0",
    }
    token_type = env.get("DUODATA_SNOWFLAKE_AUTH_TOKEN_TYPE")
    if token_type:
        headers["X-Snowflake-Authorization-Token-Type"] = token_type

    request = urllib.request.Request(
        env["DUODATA_SNOWFLAKE_ACCOUNT_URL"].rstrip("/") + "/api/v2/statements",
        data=json.dumps(payload).encode("utf-8"),
        headers=headers,
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=60) as response:
        return json.loads(response.read().decode("utf-8"))


def call_duodata_answer_metric_question(metric_name: str, question: str) -> dict[str, Any]:
    payload = {
        "metric_name": metric_name,
        "question": question,
    }
    request = urllib.request.Request(
        f"{DUODATA_API_BASE_URL}/answer-metric-question",
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json", "Accept": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=120) as response:
        return json.loads(response.read().decode("utf-8"))


def extract_month_window(question: str) -> int | None:
    match = re.search(r"last\s+(\d+)\s+months?", question.lower())
    if match:
        return int(match.group(1))
    return None


def query_workbook_metric(
    workbook_name: str,
    worksheet: str,
    date_column: str,
    value_column: str,
    aggregation: str = "avg",
    periods: int | None = None,
    filters: dict[str, Any] | None = None,
) -> dict[str, Any]:
    workbook_path = DEMO_WORKBOOKS.get(workbook_name)
    if workbook_path is None:
        raise ValueError(f"Unknown workbook: {workbook_name}")

    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook[worksheet]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {"rows": []}

    header = [str(value) for value in rows[0]]
    normalized_header = {name: idx for idx, name in enumerate(header)}
    if date_column not in normalized_header or value_column not in normalized_header:
        raise ValueError(
            f"Columns '{date_column}' and/or '{value_column}' are not present in {workbook_name}:{worksheet}"
        )

    grouped: dict[date, list[float]] = {}
    filters = filters or {}
    for row in rows[1:]:
        include = True
        for key, expected in filters.items():
            idx = normalized_header.get(key)
            if idx is None or str(row[idx]) != str(expected):
                include = False
                break
        if not include:
            continue

        raw_period = row[normalized_header[date_column]]
        raw_value = row[normalized_header[value_column]]
        if raw_period is None or raw_value is None:
            continue
        if isinstance(raw_period, datetime):
            raw_period = raw_period.date()
        grouped.setdefault(raw_period, []).append(float(raw_value))

    periods_sorted = sorted(grouped)
    if periods is not None:
        periods_sorted = periods_sorted[-periods:]

    def _aggregate(values: list[float]) -> float:
        if aggregation == "sum":
            return sum(values)
        return sum(values) / len(values)

    return {
        "rows": [
            {"period": period.isoformat(), "value": round(_aggregate(grouped[period]), 4)}
            for period in periods_sorted
        ],
        "workbook": workbook_name,
        "worksheet": worksheet,
        "valueColumn": value_column,
        "aggregation": aggregation,
        "filters": filters,
    }


async def answer_workbook_metric_question(
    bridge: MCPBridge,
    metric_name: str,
    question: str,
) -> dict[str, Any]:
    endpoint = bridge.endpoint_for_tool("get_metric")
    if endpoint is None:
        raise ValueError("DuoData get_metric tool is unavailable.")
    metric_payload, _ = await bridge.call_tool_structured(endpoint, "get_metric", {"metric_name": metric_name})
    metric = metric_payload.get("metric") if metric_payload else None
    implementations = metric_payload.get("implementations", []) if metric_payload else []
    workbook_impl = next(
        (item for item in implementations if "google drive" in str(item.get("platform") or "").lower()),
        None,
    )
    if workbook_impl is None:
        raise ValueError(f"No workbook implementation found for {metric_name}.")

    impl_name = str(workbook_impl.get("name") or "")
    parts = impl_name.split("::")
    if len(parts) != 3:
        raise ValueError(f"Unexpected workbook implementation format: {impl_name}")
    workbook_name, worksheet, value_column = parts
    periods = extract_month_window(question)
    result = await asyncio.to_thread(
        query_workbook_metric,
        workbook_name,
        worksheet,
        "reporting_month",
        value_column,
        "avg",
        periods,
        None,
    )
    return {
        "metric": metric_name,
        "question": question,
        "source": metric.get("source") if isinstance(metric, dict) else None,
        "rows": result.get("rows", []),
        "implementation": workbook_impl,
    }


def local_tool_catalog() -> list[dict[str, Any]]:
    return [
        {
            "name": "local.answer_metric_question",
            "description": (
                "Ask DuoData's execution layer to answer a Snowflake-backed metric question end to end. "
                "Use this after DuoData discovery when the metric source is Snowflake."
            ),
            "inputSchema": {
                "type": "object",
                "properties": {
                    "metric_name": {"type": "string"},
                    "question": {"type": "string"},
                },
                "required": ["metric_name", "question"],
            },
        },
        {
            "name": "local.answer_workbook_metric_question",
            "description": (
                "Ask the workbook execution layer to answer a workbook-backed metric question end to end. "
                "Use this after DuoData discovery when the metric source is the workbook."
            ),
            "inputSchema": {
                "type": "object",
                "properties": {
                    "metric_name": {"type": "string"},
                    "question": {"type": "string"},
                },
                "required": ["metric_name", "question"],
            },
        },
    ]


def build_agent_system_prompt(tool_catalog: list[dict[str, Any]]) -> str:
    tool_lines = []
    for tool in tool_catalog:
        tool_lines.append(
            f"- {tool['name']}: {tool.get('description') or 'No description'}; "
            f"input_schema={compact_json(tool.get('inputSchema') or {})}"
        )

    return (
        "You are a local analytics agent. Use tools instead of describing how data could be obtained.\n"
        "Always prefer this workflow:\n"
        "1. Use DuoData tools to discover the metric definition, source, implementations, and dimensions.\n"
        "2. Use DuoData plan_metric_execution when you need help deciding which backend should answer a metric.\n"
        "3. If the metric is Snowflake-backed, call local.answer_metric_question with the exact metric name and the original user question.\n"
        "4. If the metric is workbook-backed, call local.answer_workbook_metric_question with the exact metric name and the original user question.\n"
        "5. Do not invent raw SQL, workbook names, sheet names, or column names yourself.\n"
        "6. If the question compares multiple metrics, retrieve each relevant metric before answering.\n"
        "7. Only after you have retrieved the actual data, produce the final answer.\n\n"
        "Important rules:\n"
        "- Never answer with instructions for how the user could get the data if a tool can get it now.\n"
        "- Use the metric implementations and technical rules returned by DuoData as the authoritative query surface.\n"
        "- If the metric is not obvious, start with duodata.list_metrics and then call duodata.get_metric or duodata.resolve_metric.\n"
        "- If the user names a metric explicitly, call duodata.get_metric for that metric early.\n"
        "- Do not treat discovery metadata as the answer. Discovery must be followed by data retrieval.\n"
        "- For trend questions, retrieve the actual period-by-period values before answering.\n"
        "- If the user asks for an analytical result, trend, comparison, count, or movement, do not stop after planning. Execute the relevant local tool first.\n"
        "- When you have enough evidence, return a concise direct answer grounded in the tool results.\n\n"
        "Respond with JSON only, using one of these shapes:\n"
        '{"action":"tool","tool":"tool_name","arguments":{...}}\n'
        '{"action":"final","answer":"your grounded answer"}\n\n'
        "Example for a comparison question:\n"
        '1. {"action":"tool","tool":"duodata.get_metric","arguments":{"metric_name":"Net Sales"}}\n'
        '2. {"action":"tool","tool":"duodata.get_metric","arguments":{"metric_name":"KOL Sentiment Score"}}\n'
        '3. {"action":"tool","tool":"local.answer_metric_question","arguments":{"metric_name":"Net Sales","question":"How did Net Sales trend over the last 6 months, and did KOL Sentiment Score move in the same direction?"}}\n'
        '4. {"action":"tool","tool":"local.answer_workbook_metric_question","arguments":{"metric_name":"KOL Sentiment Score","question":"How did Net Sales trend over the last 6 months, and did KOL Sentiment Score move in the same direction?"}}\n'
        '5. {"action":"final","answer":"..."}\n\n'
        "Available tools:\n"
        + "\n".join(tool_lines)
    )


def looks_like_analysis_request(user_message: str) -> bool:
    lowered = user_message.lower()
    markers = [
        "trend",
        "how did",
        "how many",
        "what was",
        "compare",
        "moved",
        "move in the same direction",
        "last quarter",
        "last ",
        "over the ",
    ]
    return any(marker in lowered for marker in markers)


def compact_execution_result(result: dict[str, Any]) -> dict[str, Any]:
    cortex = result.get("cortexAnalystResponse") or {}
    sql_execution = cortex.get("sqlExecution") or {}
    message = cortex.get("message") or {}
    content = message.get("content") or []
    text_items = [item.get("text") for item in content if item.get("type") == "text" and item.get("text")]
    sql_items = [item.get("statement") for item in content if item.get("type") == "sql" and item.get("statement")]
    row_types = sql_execution.get("resultSetMetaData", {}).get("rowType") or []
    raw_rows = sql_execution.get("data") or []
    formatted_rows: list[dict[str, Any]] = []
    for row in raw_rows[:24]:
        formatted_row: dict[str, Any] = {}
        for idx, cell in enumerate(row):
            meta = row_types[idx] if idx < len(row_types) else {}
            name = meta.get("name") or f"col_{idx + 1}"
            cell_value: Any = cell
            if meta.get("type") == "date" and cell is not None:
                try:
                    cell_value = snowflake_date_from_serial(str(cell)).isoformat()
                except Exception:
                    cell_value = cell
            formatted_row[name] = cell_value
        formatted_rows.append(formatted_row)
    return {
        "metric": result.get("metric"),
        "question": result.get("question"),
        "summary": result.get("summary"),
        "answerable": result.get("answerable"),
        "selectedBackend": result.get("selectedBackend"),
        "backendExecution": result.get("backendExecution"),
        "analystText": text_items,
        "analystSql": sql_items[:1],
        "sqlRowType": row_types,
        "sqlData": formatted_rows,
    }


async def execute_agent_tool(
    tool_name: str,
    arguments: dict[str, Any],
    bridge: MCPBridge,
) -> dict[str, Any]:
    if tool_name == "local.run_snowflake_sql":
        statement = arguments.get("statement")
        if not statement:
            raise ValueError("local.run_snowflake_sql requires 'statement'.")
        result = await asyncio.to_thread(
            execute_snowflake_sql,
            statement,
            database=arguments.get("database"),
            schema=arguments.get("schema"),
        )
        return result

    if tool_name == "local.answer_metric_question":
        metric_name = arguments.get("metric_name")
        question = arguments.get("question")
        if not metric_name or not question:
            raise ValueError("local.answer_metric_question requires 'metric_name' and 'question'.")
        return await asyncio.to_thread(call_duodata_answer_metric_question, metric_name, question)

    if tool_name == "local.answer_workbook_metric_question":
        metric_name = arguments.get("metric_name")
        question = arguments.get("question")
        if not metric_name or not question:
            raise ValueError("local.answer_workbook_metric_question requires 'metric_name' and 'question'.")
        return await answer_workbook_metric_question(bridge, str(metric_name), str(question))

    if tool_name == "local.query_workbook_metric":
        return await asyncio.to_thread(
            query_workbook_metric,
            arguments.get("workbook_name"),
            arguments.get("worksheet"),
            arguments.get("date_column"),
            arguments.get("value_column"),
            arguments.get("aggregation", "avg"),
            arguments.get("periods"),
            arguments.get("filters"),
        )

    if tool_name.startswith("duodata."):
        bare_name = tool_name.split(".", 1)[1]
        endpoint = bridge.endpoint_for_tool(bare_name)
        if endpoint is None:
            raise ValueError(f"Unknown DuoData tool: {bare_name}")
        metric_name = arguments.get("metric_name")
        if isinstance(metric_name, str) and bare_name in {"get_metric", "resolve_metric"}:
            list_endpoint = bridge.endpoint_for_tool("list_metrics")
            if list_endpoint is not None:
                metrics_payload, _ = await bridge.call_tool_structured(list_endpoint, "list_metrics", {})
                metrics = metrics_payload.get("metrics", []) if metrics_payload else []
                wanted = normalize_metric_key(metric_name)
                for metric in metrics:
                    candidate = metric.get("name")
                    if isinstance(candidate, str) and normalize_metric_key(candidate) == wanted:
                        arguments = {**arguments, "metric_name": candidate}
                        break
        structured, text = await bridge.call_tool_structured(endpoint, bare_name, arguments)
        return structured or {"text": text}

    raise ValueError(f"Unknown tool: {tool_name}")


async def agentic_answer(
    user_message: str,
    bridge: MCPBridge,
    on_progress: Any | None = None,
) -> str | None:
    async def report_progress(message: str) -> None:
        if on_progress is None:
            return
        maybe = on_progress(message)
        if asyncio.iscoroutine(maybe):
            await maybe

    tool_catalog = [
        {
            "name": f"duodata.{tool['name']}",
            "description": tool.get("description"),
            "inputSchema": tool.get("inputSchema"),
        }
        for tool in bridge.tool_catalog()
        if tool["name"] in AGENT_DUODATA_TOOL_NAMES
    ] + local_tool_catalog()

    messages: list[dict[str, str]] = [
        {"role": "system", "content": build_agent_system_prompt(tool_catalog)},
        {"role": "user", "content": user_message},
    ]
    requires_data_execution = looks_like_analysis_request(user_message)
    executed_local_tool = False
    await report_progress("Planning tool use...")

    for _ in range(AGENT_MAX_STEPS):
        await report_progress("Thinking about the next step...")
        raw = await ollama_complete(messages, json_mode=True)
        try:
            action = parse_json_object(raw)
        except Exception:
            await report_progress("Reformatting the model response...")
            messages.append({"role": "assistant", "content": raw})
            messages.append(
                {
                    "role": "user",
                    "content": (
                        "Respond with JSON only. If this is an analytical question, do not explain how to query the data. "
                        "Choose a tool and retrieve the data first, then answer."
                    ),
                }
            )
            continue

        if action.get("action") == "final" and action.get("answer"):
            if requires_data_execution and not executed_local_tool:
                await report_progress("Need actual data before answering...")
                messages.append({"role": "assistant", "content": raw})
                messages.append(
                    {
                        "role": "user",
                        "content": (
                            "This is an analytical question. Do not stop at explanation or planning. "
                            "Use one or more local execution tools to retrieve actual results first, then return the final answer as JSON."
                        ),
                    }
                )
                continue
            return str(action["answer"])

        action_name = action.get("action")
        tool_name = action.get("tool")

        if action_name and action_name not in {"tool", "final"} and not tool_name:
            tool_name = str(action_name)

        if tool_name and not str(tool_name).startswith(("duodata.", "local.")):
            if bridge.endpoint_for_tool(str(tool_name)) is not None:
                tool_name = f"duodata.{tool_name}"

        if action_name != "tool" and not tool_name:
            messages.append({"role": "assistant", "content": raw})
            messages.append(
                {
                    "role": "user",
                    "content": "Choose a tool or return a final JSON answer. Do not answer in plain text.",
                }
            )
            continue

        if not tool_name:
            messages.append({"role": "assistant", "content": raw})
            messages.append(
                {
                    "role": "user",
                    "content": "Choose a tool or return a final JSON answer. Do not answer in plain text.",
                }
            )
            continue

        tool_name = str(tool_name)
        arguments = action.get("arguments") or {}
        if tool_name.startswith("duodata."):
            await report_progress(f"Looking up context with {tool_name.split('.', 1)[1]}...")
        elif tool_name == "local.answer_metric_question":
            await report_progress("Querying Snowflake via Cortex Analyst...")
        elif tool_name == "local.answer_workbook_metric_question":
            await report_progress("Reading workbook-backed metric data...")
        else:
            await report_progress(f"Running {tool_name}...")
        try:
            result = await execute_agent_tool(tool_name, arguments, bridge)
            if tool_name.startswith("local."):
                executed_local_tool = True
        except Exception as exc:
            result = {"error": str(exc)}
            await report_progress(f"{tool_name} returned an error. Recovering...")

        messages.append({"role": "assistant", "content": compact_json(action)})
        followup = "Continue. If you have enough information, return the final answer as JSON."
        if tool_name in {"local.answer_metric_question", "local.answer_workbook_metric_question"}:
            await report_progress("Summarizing the retrieved results...")
            followup = (
                "You now have actual source-grounded result data. "
                "Return the final answer as JSON unless you still need to retrieve a second metric for a comparison question."
            )
        messages.append(
            {
                "role": "user",
                "content": (
                    f"Tool result for {tool_name}:\n"
                    f"{json.dumps(result, ensure_ascii=True, indent=2)[:12000]}\n\n"
                    + followup
                ),
            }
        )

    return None


@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="index.html",
        context={"conversation_id": str(uuid4())},
    )


@app.get("/api/status")
async def status() -> JSONResponse:
    available, detail = await ollama_status()
    return JSONResponse(content={"ok": bool(available), "detail": detail, "mcp_available": mcp_types is not None})


@app.post("/api/mcp/inspect")
async def inspect_mcp(payload: MCPInspectRequest) -> JSONResponse:
    bridge = await MCPBridge.create(payload.endpoints)
    try:
        return JSONResponse(content={"ok": True, "servers": bridge.summaries()})
    finally:
        await bridge.close()


@app.post("/api/chat/stream")
async def chat_stream(payload: ChatRequest):
    conversation_id = payload.conversation_id or str(uuid4())
    managed = await store.get_or_create(
        conversation_id,
        payload.system_prompt,
        payload.mcp_endpoints,
    )

    async def event_stream():
        meta: dict[str, Any] = {"type": "meta", "conversation_id": conversation_id}
        if managed.mcp_bridge is not None:
            meta["mcp_servers"] = managed.mcp_bridge.summaries()
        yield f"data: {json.dumps(meta)}\n\n"
        try:
            if managed.mcp_bridge is not None:
                progress_queue: asyncio.Queue[str] = asyncio.Queue()

                async def on_progress(message: str) -> None:
                    await progress_queue.put(message)

                task = asyncio.create_task(agentic_answer(payload.message, managed.mcp_bridge, on_progress=on_progress))
                while True:
                    if task.done() and progress_queue.empty():
                        break
                    try:
                        progress = await asyncio.wait_for(progress_queue.get(), timeout=0.25)
                        yield f"data: {json.dumps({'type': 'progress', 'content': progress})}\n\n"
                    except asyncio.TimeoutError:
                        pass
                analytics_answer = await task
            else:
                analytics_answer = None
            if analytics_answer:
                yield f"data: {json.dumps({'type': 'chunk', 'content': analytics_answer})}\n\n"
                managed.history.append({"role": "user", "content": payload.message})
                managed.history.append({"role": "assistant", "content": analytics_answer})
                yield f"data: {json.dumps({'type': 'done'})}\n\n"
                return

            prompt = payload.message
            if managed.mcp_bridge is not None:
                grounding = await managed.mcp_bridge.build_grounding_context(payload.message)
                if grounding:
                    prompt = (
                        "Use the following MCP-grounded business context when answering. "
                        "Treat it as authoritative source context from DuoData.\n\n"
                        f"{grounding}\n\n"
                        f"User question: {payload.message}"
                    )
            messages: list[dict[str, str]] = []
            if payload.system_prompt:
                messages.append({"role": "system", "content": payload.system_prompt})
            messages.extend(managed.history)
            messages.append({"role": "user", "content": prompt})

            response_text = ""
            async for event in ollama_stream_chat(messages):
                chunk = event.get("message", {}).get("content", "")
                if chunk:
                    response_text += chunk
                    yield f"data: {json.dumps({'type': 'chunk', 'content': response_text})}\n\n"

                if event.get("done"):
                    managed.history.append({"role": "user", "content": payload.message})
                    managed.history.append({"role": "assistant", "content": response_text})
            yield f"data: {json.dumps({'type': 'done'})}\n\n"
        except Exception as exc:  # pragma: no cover
            yield f"data: {json.dumps({'type': 'error', 'content': str(exc)})}\n\n"

    return StreamingResponse(event_stream(), media_type="text/event-stream")


@app.post("/api/chat/reset")
async def reset_chat(payload: ResetRequest) -> JSONResponse:
    removed = await store.reset(payload.conversation_id)
    return JSONResponse(content={"ok": True, "removed": removed})


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="127.0.0.1", port=8000)
